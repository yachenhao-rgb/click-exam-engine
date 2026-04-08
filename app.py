# -*- coding: utf-8 -*-
"""
Flask Web应用入口
智能出题引擎Web版
"""

import os
import json
import uuid
import re
from datetime import datetime
from typing import Dict, List, Optional, Any
from flask import Flask, render_template, request, jsonify, send_file, flash, redirect, url_for
from flask_login import LoginManager, login_user, logout_user, login_required, current_user

# 创建Flask应用
app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY')
if not app.secret_key:
    # 开发环境使用默认密钥，生产环境必须设置
    app.secret_key = 'dev-secret-key-change-in-production'

# Flask-Login配置
login_manager = LoginManager()
login_manager.login_view = 'login'  # 未登录时重定向到登录页面
login_manager.login_message = '请先登录'
login_manager.init_app(app)

# 导入用户模型
from models.user import User, UserHistory

@login_manager.user_loader
def load_user(user_id):
    """Flask-Login用户加载器"""
    return User.get_by_id(DB_PATH, int(user_id))
app.config['UPLOAD_FOLDER'] = 'output'
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 最大16MB

# 确保输出目录存在
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
os.makedirs('data', exist_ok=True)
os.makedirs('data/知识库数据', exist_ok=True)

# 导入核心模块
from core.exam_generator import ExamGenerator
from core.reading_generator import ReadingGenerator, generate_reading_questions
from core.true_false_generator import TrueFalseGenerator, generate_true_false_questions
from core.open_question_generator import OpenQuestionGenerator, generate_open_questions
from core.word_matching_generator import WordMatchingGenerator, generate_word_matching
from core.kb_manager import KnowledgeBaseManager
from core.kb_backup import KnowledgeBaseBackup
from core.kb_importer import KnowledgeBaseImporter, create_template_excel
from core.kb_validator import KnowledgeBaseValidator
from core.docx_exporter import export_to_word
from core.pdf_exporter import export_to_pdf
from core.kb_exporter import KnowledgeBaseExporter
from core.kb_statistics import KnowledgeBaseStatistics
from core.uploader import process_uploaded_file, validate_file, get_file_info
from core.cefr_analyzer import CEFRAnalyzer
from core.lightweight_tools import (
    generate_discussion_questions,
    generate_sentences,
    generate_dialogue,
    generate_error_correction,
    generate_essential_vocabulary,
    generate_pros_cons,
    generate_famous_quotes,
    generate_essay_topics,
    generate_interesting_facts,
    generate_three_titles,
    generate_creative_writing,
    generate_odd_one_out
)

# 全局变量
DB_PATH = 'data/knowledge_base.db'
KB_DATA_PATH = 'data/知识库数据'
BACKUP_DIR = 'data/backups'


def get_kb_manager():
    """获取知识库管理器实例"""
    return KnowledgeBaseManager(DB_PATH)


def get_backup_manager():
    """获取备份管理器实例"""
    return KnowledgeBaseBackup(DB_PATH, BACKUP_DIR)


def get_generator():
    """获取题目生成器实例"""
    return ExamGenerator(DB_PATH)


@app.route('/')
def index():
    """首页/出题页面"""
    return render_template('index.html')


@app.route('/kb')
def kb_manage():
    """知识库管理页面"""
    return render_template('kb_manage.html')


@app.route('/history')
def history():
    """历史记录页面"""
    history_files = []
    output_dir = app.config['UPLOAD_FOLDER']
    
    if os.path.exists(output_dir):
        for f in os.listdir(output_dir):
            if f.endswith('.json'):
                filepath = os.path.join(output_dir, f)
                stats = os.stat(filepath)
                history_files.append({
                    'filename': f,
                    'size': stats.st_size,
                    'created': datetime.fromtimestamp(stats.st_ctime).strftime('%Y-%m-%d %H:%M:%S'),
                    'task_id': f.replace('.json', '')
                })
    
    # 按时间倒序
    history_files.sort(key=lambda x: x['created'], reverse=True)
    return render_template('history.html', history=history_files)


@app.route('/profile')
def profile():
    """个人中心页面"""
    if current_user.is_authenticated:
        return render_template('profile.html')
    else:
        flash('请先登录', 'warning')
        return redirect(url_for('index'))


# ==================== 用户认证接口 ====================

def validate_email(email):
    """验证邮箱格式"""
    pattern = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
    return re.match(pattern, email) is not None


def validate_username(username):
    """验证用户名格式"""
    if len(username) < 3 or len(username) > 20:
        return False
    pattern = r'^[a-zA-Z0-9_\u4e00-\u9fa5]+$'
    return re.match(pattern, username) is not None


@app.route('/register', methods=['POST'])
def register():
    """
    用户注册接口
    
    输入：username, email, password
    输出：注册结果JSON
    """
    try:
        data = request.get_json()
        
        username = data.get('username', '').strip()
        email = data.get('email', '').strip().lower()
        password = data.get('password', '')
        
        # 参数验证
        if not username or not email or not password:
            return jsonify({'success': False, 'error': '请填写所有必填项'}), 400
        
        if not validate_username(username):
            return jsonify({
                'success': False, 
                'error': '用户名长度3-20位，只能包含字母、数字、下划线和中文'
            }), 400
        
        if not validate_email(email):
            return jsonify({'success': False, 'error': '请输入有效的邮箱地址'}), 400
        
        if len(password) < 6:
            return jsonify({'success': False, 'error': '密码长度至少6位'}), 400
        
        # 创建用户
        success, message, user = User.create(DB_PATH, username, email, password)
        
        if success:
            # 自动登录
            login_user(user)
            return jsonify({
                'success': True,
                'message': '注册成功',
                'user': user.to_dict()
            })
        else:
            return jsonify({'success': False, 'error': message}), 400
            
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/login', methods=['POST'])
def login():
    """
    用户登录接口
    
    输入：username_or_email, password
    输出：登录结果JSON
    """
    try:
        data = request.get_json()
        
        username_or_email = data.get('username_or_email', '').strip()
        password = data.get('password', '')
        
        if not username_or_email or not password:
            return jsonify({'success': False, 'error': '请填写用户名和密码'}), 400
        
        # 根据输入格式判断是用户名还是邮箱
        if '@' in username_or_email:
            user = User.get_by_email(DB_PATH, username_or_email.lower())
        else:
            user = User.get_by_username(DB_PATH, username_or_email)
        
        if user and user.check_password(password):
            # 更新最后登录时间
            user.update_last_login(DB_PATH)
            # 登录
            login_user(user, remember=True)
            
            return jsonify({
                'success': True,
                'message': '登录成功',
                'user': user.to_dict()
            })
        else:
            return jsonify({'success': False, 'error': '用户名或密码错误'}), 401
            
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/logout', methods=['POST'])
@login_required
def logout():
    """
    用户登出接口
    
    输出：登出结果JSON
    """
    try:
        logout_user()
        return jsonify({'success': True, 'message': '已退出登录'})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/user/profile')
def api_user_profile():
    """
    获取当前用户信息接口
    
    输出：用户信息JSON
    """
    try:
        if current_user.is_authenticated:
            return jsonify({
                'success': True,
                'user': current_user.to_dict(),
                'is_authenticated': True
            })
        else:
            return jsonify({
                'success': True,
                'is_authenticated': False,
                'user': None
            })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/user/history')
def api_user_history():
    """
    获取当前用户历史记录接口
    
    输出：历史记录列表JSON
    """
    try:
        if not current_user.is_authenticated:
            return jsonify({'success': False, 'error': '请先登录'}), 401
        
        limit = request.args.get('limit', 50, type=int)
        history = UserHistory.get_by_user(DB_PATH, current_user.id, limit)
        
        return jsonify({
            'success': True,
            'history': history,
            'count': len(history)
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/user/check-session')
def api_check_session():
    """
    检查用户登录状态
    
    输出：登录状态JSON
    """
    try:
        if current_user.is_authenticated:
            return jsonify({
                'success': True,
                'is_authenticated': True,
                'user': current_user.to_dict()
            })
        else:
            return jsonify({
                'success': True,
                'is_authenticated': False,
                'user': None
            })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


# ==================== API接口 ====================

@app.route('/api/generate', methods=['POST'])
def api_generate():
    """
    生成题目接口
    
    输入：text, question_type, difficulty, num_blanks
    输出：生成的题目JSON
    """
    try:
        data = request.get_json()
        
        text = data.get('text', '').strip()
        question_type = data.get('question_type', 'cloze')
        difficulty = data.get('difficulty', 'medium')
        num_blanks = int(data.get('num_blanks', 10))
        
        if not text:
            return jsonify({'success': False, 'error': '请输入文本内容'}), 400
        
        if len(text) < 50:
            return jsonify({'success': False, 'error': '文本内容太短，至少需要50个字符'}), 400
        
        # 根据题型调用不同的生成器
        if question_type == 'cloze':
            # 完形填空
            generator = get_generator()
            result = generator.generate_exam(text, num_blanks)
            
        elif question_type == 'reading':
            # 阅读理解
            reading_gen = ReadingGenerator()
            raw_result = reading_gen.generate_questions(text, num_questions=num_blanks, difficulty=difficulty)
            result = reading_gen._generate_reading_exam_output(raw_result)
                
        elif question_type == 'grammar':
            # 语法填空 - 检查是否有专门的语法填空生成器
            try:
                from core.grammar_generator import GrammarGenerator
                grammar_gen = GrammarGenerator(DB_PATH)
                result = grammar_gen.generate_questions(text, num_blanks, difficulty)
            except ImportError:
                # 如果还没有实现，返回提示信息
                result = {
                    'text_with_blanks': text,
                    'questions': [{
                        'blank_number': 1,
                        'answer': '语法填空',
                        'hint': '模块开发中',
                        'explanation': 'V1.0版本正在开发语法填空功能'
                    }],
                    'question_type': 'grammar'
                }
        elif question_type == 'true_false':
            # 判断题
            tf_gen = TrueFalseGenerator()
            result = tf_gen.generate_true_false(text, num_questions=num_blanks, difficulty=difficulty)
        elif question_type == 'open_questions':
            # 开放问答题
            oq_gen = OpenQuestionGenerator()
            language = data.get('language', 'en')
            result = oq_gen.generate_open_questions(text, num_questions=num_blanks, difficulty=difficulty, language=language)
        elif question_type == 'word_matching':
            # 词汇匹配题
            wm_gen = WordMatchingGenerator()
            language = data.get('language', 'en')
            result = wm_gen.generate_word_matching(text=text, num_pairs=num_blanks, language=language)
        else:
            return jsonify({'success': False, 'error': '不支持的题型'}), 400
        
        # 生成任务ID
        task_id = str(uuid.uuid4())[:8]
        
        # 保存到文件历史记录
        history_data = {
            'task_id': task_id,
            'created_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            'question_type': question_type,
            'difficulty': difficulty,
            'num_questions': num_blanks,
            'original_text': text[:200] + '...' if len(text) > 200 else text,
            'result': result
        }
        
        history_path = os.path.join(app.config['UPLOAD_FOLDER'], f'{task_id}.json')
        with open(history_path, 'w', encoding='utf-8') as f:
            json.dump(history_data, f, ensure_ascii=False, indent=2)
        
        # 如果用户已登录，保存到数据库
        if current_user.is_authenticated:
            try:
                UserHistory.create(
                    db_path=DB_PATH,
                    user_id=current_user.id,
                    task_id=task_id,
                    question_type=question_type,
                    difficulty=difficulty,
                    num_questions=num_blanks,
                    original_text=text[:200] + '...' if len(text) > 200 else text
                )
            except Exception as history_error:
                # 历史记录保存失败不影响主流程
                print(f"保存用户历史记录失败: {history_error}")
        
        # 返回结果（兼容前端格式）
        response = {
            'success': True,
            'task_id': task_id,
            'questions': result.get('questions', []),
            'question_type': question_type
        }
        
        # 根据题型添加不同字段
        if question_type == 'reading':
            response['article'] = result.get('article', text)
            response['article_title'] = result.get('article_title', '')
            response['article_type'] = result.get('article_type', '')
            response['word_count'] = result.get('word_count', 0)
            response['total_questions'] = result.get('total_questions', len(result.get('questions', [])))
        elif question_type == 'true_false':
            # 判断题 - 返回原文
            response['original_text'] = result.get('original_text', text)
            response['total_questions'] = result.get('total_questions', len(result.get('questions', [])))
            response['true_count'] = result.get('true_count', 0)
            response['false_count'] = result.get('false_count', 0)
        elif question_type == 'open_questions':
            # 开放问答题 - 返回原文和主题
            response['original_text'] = result.get('original_text', text)
            response['main_topic'] = result.get('main_topic', '')
            response['total_questions'] = result.get('total_questions', len(result.get('questions', [])))
            response['language'] = result.get('language', 'en')
        elif question_type == 'word_matching':
            # 词汇匹配题 - 返回配对和选项
            response['original_text'] = result.get('original_text', text)
            response['total_pairs'] = result.get('total_pairs', 0)
            response['language'] = result.get('language', 'en')
            response['items'] = result.get('items', [])
            response['options'] = result.get('options', [])
            response['answers'] = result.get('answers', [])
        else:
            # 完形填空/语法填空 - 返回原文和带空文本
            response['original_text'] = result.get('original_text', text)
            response['text_with_blanks'] = result.get('blanked_text', result.get('text_with_blanks', text))
        
        return jsonify(response)
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/export/<task_id>')
def api_export(task_id):
    """
    导出Word文件接口
    
    输出：Word文件下载
    """
    try:
        # 读取历史记录
        history_path = os.path.join(app.config['UPLOAD_FOLDER'], f'{task_id}.json')
        
        if not os.path.exists(history_path):
            return jsonify({'success': False, 'error': '任务不存在'}), 404
        
        with open(history_path, 'r', encoding='utf-8') as f:
            history_data = json.load(f)
        
        result = history_data['result']
        
        # 导出Word
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        output_filename = f'试卷_{task_id}_{timestamp}.docx'
        output_path = os.path.join(app.config['UPLOAD_FOLDER'], output_filename)
        
        export_to_word(result, output_path)
        
        return send_file(
            output_path,
            as_attachment=True,
            download_name=output_filename,
            mimetype='application/vnd.openxmlformats-officedocument.wordprocessingml.document'
        )
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/export/pdf/<task_id>')
def api_export_pdf(task_id):
    """
    导出PDF文件接口
    
    输出：PDF文件下载
    """
    try:
        # 读取历史记录
        history_path = os.path.join(app.config['UPLOAD_FOLDER'], f'{task_id}.json')
        
        if not os.path.exists(history_path):
            return jsonify({'success': False, 'error': '任务不存在'}), 404
        
        with open(history_path, 'r', encoding='utf-8') as f:
            history_data = json.load(f)
        
        result = history_data['result']
        
        # 导出PDF
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        output_filename = f'试卷_{task_id}_{timestamp}.pdf'
        output_path = os.path.join(app.config['UPLOAD_FOLDER'], output_filename)
        
        # 检查是否包含答案
        include_answers = request.args.get('include_answers', 'true').lower() == 'true'
        
        export_to_pdf(result, output_path, include_answers)
        
        return send_file(
            output_path,
            as_attachment=True,
            download_name=output_filename,
            mimetype='application/pdf'
        )
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/export/html/<task_id>')
def api_export_html(task_id):
    """
    导出打印友好HTML接口
    
    输出：HTML内容（用于打印或复制）
    """
    try:
        # 读取历史记录
        history_path = os.path.join(app.config['UPLOAD_FOLDER'], f'{task_id}.json')
        
        if not os.path.exists(history_path):
            return jsonify({'success': False, 'error': '任务不存在'}), 404
        
        with open(history_path, 'r', encoding='utf-8') as f:
            history_data = json.load(f)
        
        result = history_data['result']
        question_type = history_data.get('question_type', 'cloze')
        
        # 生成HTML内容
        html_content = _generate_print_html(result, question_type)
        
        return jsonify({
            'success': True,
            'html': html_content,
            'text': _generate_plain_text(result, question_type)
        })
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


def _generate_print_html(exam_data: Dict, question_type: str) -> str:
    """生成打印友好的HTML内容"""
    html_parts = []
    html_parts.append('<div class="print-exam">')
    
    # 根据题型生成不同内容
    if question_type == 'cloze':
        html_parts.append('<h1 class="print-title">英语完形填空试卷</h1>')
        total = exam_data.get('total_blanks', len(exam_data.get('questions', [])))
        html_parts.append(f'<p class="print-info">共{total}小题，每小题1分，满分{total}分</p>')
        html_parts.append('<p class="print-instruction">阅读下面短文，从短文后各题所给的A、B、C、D四个选项中，选出最佳选项。</p>')
        
        blanked_text = exam_data.get('blanked_text', exam_data.get('text_with_blanks', ''))
        if blanked_text:
            html_parts.append(f'<div class="print-text">{blanked_text.replace(chr(10), "<br>")}</div>')
        
        html_parts.append('<h2 class="print-section">【题目】</h2>')
        html_parts.append('<div class="print-questions">')
        for q in exam_data.get('questions', []):
            opts = ''.join([f'<span class="option">{chr(65+j)}. {opt}</span>' for j, opt in enumerate(q.get('options', []))])
            html_parts.append(f'<div class="question"><span class="q-num">{q["number"]}.</span> {opts}</div>')
        html_parts.append('</div>')
        
    elif question_type == 'reading':
        html_parts.append('<h1 class="print-title">英语阅读理解试卷</h1>')
        article_title = exam_data.get('article_title', '')
        if article_title:
            html_parts.append(f'<p class="print-info"><b>文章标题：</b>{article_title}</p>')
        article = exam_data.get('article', exam_data.get('original_text', ''))
        if article:
            html_parts.append(f'<div class="print-text">{article.replace(chr(10), "<br>")}</div>')
        html_parts.append('<h2 class="print-section">【阅读题目】</h2>')
        for q in exam_data.get('questions', []):
            opts = ''.join([f'<div class="option">{chr(65+j)}. {opt}</div>' for j, opt in enumerate(q.get('options', []))])
            html_parts.append(f'<div class="question"><b>{q["number"]}.</b> {q.get("question", q.get("text", ""))}<div class="options">{opts}</div></div>')
    
    elif question_type == 'true_false':
        html_parts.append('<h1 class="print-title">英语判断题试卷</h1>')
        total = exam_data.get('total_questions', len(exam_data.get('questions', [])))
        html_parts.append(f'<p class="print-info">共{total}小题</p>')
        html_parts.append('<p class="print-instruction">阅读下面句子，判断句子内容是否正确。正确的写True，错误的写False。</p>')
        html_parts.append('<h2 class="print-section">【判断题】</h2>')
        for q in exam_data.get('questions', []):
            html_parts.append(f'<div class="question"><b>{q["number"]}.</b> {q.get("statement", q.get("text", ""))} (     )</div>')
    
    elif question_type == 'open_questions':
        html_parts.append('<h1 class="print-title">英语开放问答题试卷</h1>')
        html_parts.append('<p class="print-instruction">请根据文章内容回答以下问题。</p>')
        original_text = exam_data.get('original_text', '')
        if original_text:
            html_parts.append(f'<div class="print-text"><b>原文：</b><br>{original_text.replace(chr(10), "<br>")}</div>')
        html_parts.append('<h2 class="print-section">【开放问答题】</h2>')
        for q in exam_data.get('questions', []):
            html_parts.append(f'<div class="question"><b>{q["number"]}.</b> {q.get("question", q.get("text", ""))}</div>')
    
    elif question_type == 'word_matching':
        html_parts.append('<h1 class="print-title">英语词汇匹配题试卷</h1>')
        html_parts.append('<p class="print-instruction">请将下列单词与正确的释义匹配。</p>')
        html_parts.append('<h2 class="print-section">【词汇】</h2>')
        for item in exam_data.get('items', []):
            html_parts.append(f'<div class="question"><b>{item["number"]}.</b> {item["word"]}  →  (    )</div>')
        html_parts.append('<h2 class="print-section">【释义】</h2>')
        for opt in exam_data.get('options', []):
            html_parts.append(f'<div class="option"><b>{opt["letter"]}.</b> {opt["definition"]}</div>')
    
    html_parts.append('</div>')
    return ''.join(html_parts)


def _generate_plain_text(exam_data: Dict, question_type: str) -> str:
    """生成纯文本内容（用于复制）"""
    text_parts = []
    
    if question_type == 'cloze':
        text_parts.append('英语完形填空试卷')
        text_parts.append('=' * 40)
        blanked_text = exam_data.get('blanked_text', exam_data.get('text_with_blanks', ''))
        if blanked_text:
            text_parts.append(blanked_text)
        text_parts.append('')
        text_parts.append('【题目】')
        for q in exam_data.get('questions', []):
            opts = '  '.join([f'{chr(65+j)}. {opt}' for j, opt in enumerate(q.get('options', []))])
            text_parts.append(f'{q["number"]}. {opts}')
    
    elif question_type == 'reading':
        text_parts.append('英语阅读理解试卷')
        text_parts.append('=' * 40)
        article = exam_data.get('article', exam_data.get('original_text', ''))
        if article:
            text_parts.append(article)
        text_parts.append('')
        text_parts.append('【阅读题目】')
        for q in exam_data.get('questions', []):
            text_parts.append(f'{q["number"]}. {q.get("question", q.get("text", ""))}')
            for j, opt in enumerate(q.get('options', [])):
                text_parts.append(f'   {chr(65+j)}. {opt}')
    
    elif question_type == 'true_false':
        text_parts.append('英语判断题试卷')
        text_parts.append('=' * 40)
        for q in exam_data.get('questions', []):
            text_parts.append(f'{q["number"]}. {q.get("statement", q.get("text", ""))} (     )')
    
    elif question_type == 'open_questions':
        text_parts.append('英语开放问答题试卷')
        text_parts.append('=' * 40)
        for q in exam_data.get('questions', []):
            text_parts.append(f'{q["number"]}. {q.get("question", q.get("text", ""))}')
    
    elif question_type == 'word_matching':
        text_parts.append('英语词汇匹配题试卷')
        text_parts.append('=' * 40)
        text_parts.append('【词汇】')
        for item in exam_data.get('items', []):
            text_parts.append(f'{item["number"]}. {item["word"]}  →  (    )')
        text_parts.append('')
        text_parts.append('【释义】')
        for opt in exam_data.get('options', []):
            text_parts.append(f'{opt["letter"]}. {opt["definition"]}')
    
    return '\n'.join(text_parts)


@app.route('/api/export/answers/<task_id>')
def api_export_answers(task_id):
    """
    导出答案接口
    
    输出：答案文本
    """
    try:
        # 读取历史记录
        history_path = os.path.join(app.config['UPLOAD_FOLDER'], f'{task_id}.json')
        
        if not os.path.exists(history_path):
            return jsonify({'success': False, 'error': '任务不存在'}), 404
        
        with open(history_path, 'r', encoding='utf-8') as f:
            history_data = json.load(f)
        
        result = history_data['result']
        question_type = history_data.get('question_type', 'cloze')
        
        # 生成答案文本
        answer_text = _generate_answer_text(result, question_type)
        
        return jsonify({
            'success': True,
            'text': answer_text
        })
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


def _generate_answer_text(exam_data: Dict, question_type: str) -> str:
    """生成答案文本"""
    text_parts = []
    text_parts.append('参考答案')
    text_parts.append('=' * 40)
    
    questions = exam_data.get('questions', [])
    
    if question_type == 'cloze':
        for q in questions:
            answer = q.get('correct_letter', q.get('answer', ''))
            text_parts.append(f'{q["number"]}. {answer}')
        text_parts.append('')
        text_parts.append('答案解析：')
        for q in questions:
            if q.get('explanation'):
                text_parts.append(f'{q["number"]}. {q["explanation"]}')
    
    elif question_type == 'reading':
        for q in questions:
            answer = q.get('correct_letter', q.get('answer', ''))
            text_parts.append(f'{q["number"]}. {answer}')
    
    elif question_type == 'true_false':
        for q in questions:
            answer = q.get('answer', q.get('correct_letter', ''))
            text_parts.append(f'{q["number"]}. {answer}')
    
    elif question_type == 'open_questions':
        for q in questions:
            answer = q.get('sample_answer', q.get('answer', ''))
            if answer:
                text_parts.append(f'{q["number"]}. {answer}')
            else:
                text_parts.append(f'{q["number"]}. (参考答案待添加)')
    
    elif question_type == 'word_matching':
        answers = exam_data.get('answers', [])
        for i, ans in enumerate(answers, 1):
            text_parts.append(f'{i}. {ans}')
    
    return '\n'.join(text_parts)
def api_kb_stats():
    """
    获取知识库统计信息
    
    输出：知识库统计信息
    """
    try:
        kb = get_kb_manager()
        kb.connect()
        
        cursor = kb.conn.cursor()
        
        # 词汇数
        cursor.execute("SELECT COUNT(*) FROM word_frequency")
        vocab_count = cursor.fetchone()[0]
        
        # 陷阱数
        cursor.execute("SELECT COUNT(*) FROM trap_library")
        trap_count = cursor.fetchone()[0]
        
        # 真题示例数
        cursor.execute("SELECT COUNT(*) FROM example_questions")
        example_count = cursor.fetchone()[0]
        
        kb.close()
        
        return jsonify({
            'success': True,
            'stats': {
                'vocab_count': vocab_count,
                'trap_count': trap_count,
                'example_count': example_count
            }
        })
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/kb/validate')
def api_kb_validate():
    """
    知识库数据校验接口
    
    功能：执行完整的数据校验，包括重复检测、格式校验、冲突检测
    输出：校验报告
    """
    try:
        # 创建校验器并执行校验
        validator = KnowledgeBaseValidator(DB_PATH)
        report = validator.validate_all()
        
        return jsonify({
            'success': True,
            'report': report
        })
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/kb/validate/report')
def api_kb_validate_report():
    """
    获取文本格式的校验报告
    
    输出：文本格式的校验报告
    """
    try:
        validator = KnowledgeBaseValidator(DB_PATH)
        validator.connect()
        report = validator.validate_all()
        text_report = validator._format_text_report()
        validator.close()
        
        return jsonify({
            'success': True,
            'report': text_report
        })
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/kb/validate/fixes')
def api_kb_validate_fixes():
    """
    获取可修复的问题列表
    
    输出：可自动修复和需要手动修复的问题列表
    """
    try:
        validator = KnowledgeBaseValidator(DB_PATH)
        validator.connect()
        report = validator.validate_all()
        fixes = validator.get_fixes()
        validator.close()
        
        return jsonify({
            'success': True,
            'fixes': fixes
        })
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/kb/validate/fix/<fix_type>', methods=['POST'])
def api_kb_validate_fix(fix_type):
    """
    执行数据修复
    
    参数：
    - fix_type: 修复类型 (fix_duplicates, fix_format, etc.)
    输入：修复参数
    """
    try:
        data = request.get_json() or {}
        validator = KnowledgeBaseValidator(DB_PATH)
        validator.connect()
        
        result = {'success': True, 'fixed_count': 0, 'errors': []}
        
        if fix_type == 'fix_duplicates':
            # 修复重复数据
            cursor = validator.conn.cursor()
            
            # 处理词汇库重复
            vocab_dups = validator.report['summary']['duplicates']['vocab']
            for dup in vocab_dups:
                # 保留第一条，删除其他
                ids = [e['id'] for e in dup['entries']]
                if len(ids) > 1:
                    cursor.execute("""
                        DELETE FROM word_frequency 
                        WHERE id IN ({})
                    """.format(','.join('?' * len(ids[1:]))), ids[1:])
                    result['fixed_count'] += len(ids) - 1
            
            # 处理陷阱库重复
            trap_dups = validator.report['summary']['duplicates']['traps']
            for dup in trap_dups:
                ids = [e['id'] for e in dup['entries']]
                if len(ids) > 1:
                    cursor.execute("""
                        DELETE FROM trap_library 
                        WHERE id IN ({})
                    """.format(','.join('?' * len(ids[1:]))), ids[1:])
                    result['fixed_count'] += len(ids) - 1
            
            validator.conn.commit()
            
        elif fix_type == 'fix_format':
            # 修复格式错误
            cursor = validator.conn.cursor()
            
            # 修复词性标签（无效的设为空）
            pos_errors = validator.report['summary']['format_errors']['vocab']['pos_errors']
            for error in pos_errors:
                cursor.execute("UPDATE word_frequency SET pos_tag = NULL WHERE id = ?", (error['id'],))
                result['fixed_count'] += 1
            
            # 标准化词性标签（警告的进行转换）
            pos_warnings = validator.report['summary']['format_errors']['vocab']['pos_warnings']
            for warning in pos_warnings:
                normalized = KnowledgeBaseValidator.POS_TAG_NORMALIZE.get(warning['pos_tag'].lower())
                if normalized:
                    cursor.execute("UPDATE word_frequency SET pos_tag = ? WHERE id = ?", 
                                  (normalized, warning['id']))
                    result['fixed_count'] += 1
            
            # 修复频率等级（无效的设为medium）
            freq_errors = validator.report['summary']['format_errors']['vocab']['freq_errors']
            for error in freq_errors:
                cursor.execute("UPDATE word_frequency SET frequency_level = 'medium' WHERE id = ?", (error['id'],))
                result['fixed_count'] += 1
            
            # 标准化频率等级（警告的进行转换）
            freq_warnings = validator.report['summary']['format_errors']['vocab']['freq_warnings']
            for warning in freq_warnings:
                normalized = KnowledgeBaseValidator.FREQ_LEVEL_NORMALIZE.get(warning['frequency_level'].lower())
                if normalized:
                    cursor.execute("UPDATE word_frequency SET frequency_level = ? WHERE id = ?", 
                                  (normalized, warning['id']))
                    result['fixed_count'] += 1
            
            # 修复陷阱类型（无效的设为词汇辨析）
            type_errors = validator.report['summary']['format_errors']['traps']['type_errors']
            for error in type_errors:
                cursor.execute("UPDATE trap_library SET trap_type = '词汇辨析' WHERE id = ?", (error['id'],))
                result['fixed_count'] += 1
            
            # 标准化陷阱类型（警告的进行转换）
            type_warnings = validator.report['summary']['format_errors']['traps']['type_warnings']
            for warning in type_warnings:
                normalized = KnowledgeBaseValidator.TRAP_TYPE_NORMALIZE.get(warning['trap_type'])
                if normalized:
                    cursor.execute("UPDATE trap_library SET trap_type = ? WHERE id = ?", 
                                  (normalized, warning['id']))
                    result['fixed_count'] += 1
            
            validator.conn.commit()
            
        elif fix_type == 'fix_missing':
            # 修复缺失数据
            cursor = validator.conn.cursor()
            
            # 为缺失频率等级的词汇设置默认值
            cursor.execute("""
                UPDATE word_frequency 
                SET frequency_level = 'medium' 
                WHERE frequency_level IS NULL OR frequency_level = ''
            """)
            result['fixed_count'] = cursor.rowcount
            
            # 为缺失解释的陷阱添加默认解释
            cursor.execute("""
                UPDATE trap_library 
                SET explanation = '请补充解释说明' 
                WHERE explanation IS NULL OR explanation = ''
            """)
            result['fixed_count'] += cursor.rowcount
            
            validator.conn.commit()
            
        elif fix_type == 'fix_all':
            # 执行所有修复
            cursor = validator.conn.cursor()
            
            # 1. 修复重复
            vocab_dups = validator.report['summary']['duplicates']['vocab']
            for dup in vocab_dups:
                ids = [e['id'] for e in dup['entries']]
                if len(ids) > 1:
                    cursor.execute(f"""
                        DELETE FROM word_frequency 
                        WHERE id IN ({','.join('?' * len(ids[1:]))})
                    """, ids[1:])
                    result['fixed_count'] += len(ids) - 1
            
            trap_dups = validator.report['summary']['duplicates']['traps']
            for dup in trap_dups:
                ids = [e['id'] for e in dup['entries']]
                if len(ids) > 1:
                    cursor.execute(f"""
                        DELETE FROM trap_library 
                        WHERE id IN ({','.join('?' * len(ids[1:]))})
                    """, ids[1:])
                    result['fixed_count'] += len(ids) - 1
            
            # 2. 修复格式（错误）
            format_errors = validator.report['summary']['format_errors']
            
            for error in format_errors['vocab']['pos_errors']:
                cursor.execute("UPDATE word_frequency SET pos_tag = NULL WHERE id = ?", (error['id'],))
                result['fixed_count'] += 1
            
            for error in format_errors['vocab']['freq_errors']:
                cursor.execute("UPDATE word_frequency SET frequency_level = 'medium' WHERE id = ?", (error['id'],))
                result['fixed_count'] += 1
            
            for error in format_errors['traps']['type_errors']:
                cursor.execute("UPDATE trap_library SET trap_type = '词汇辨析' WHERE id = ?", (error['id'],))
                result['fixed_count'] += 1
            
            # 3. 标准化格式（警告）
            for warning in format_errors['vocab']['pos_warnings']:
                normalized = KnowledgeBaseValidator.POS_TAG_NORMALIZE.get(warning['pos_tag'].lower())
                if normalized:
                    cursor.execute("UPDATE word_frequency SET pos_tag = ? WHERE id = ?", 
                                  (normalized, warning['id']))
                    result['fixed_count'] += 1
            
            for warning in format_errors['vocab']['freq_warnings']:
                normalized = KnowledgeBaseValidator.FREQ_LEVEL_NORMALIZE.get(warning['frequency_level'].lower())
                if normalized:
                    cursor.execute("UPDATE word_frequency SET frequency_level = ? WHERE id = ?", 
                                  (normalized, warning['id']))
                    result['fixed_count'] += 1
            
            for warning in format_errors['traps']['type_warnings']:
                normalized = KnowledgeBaseValidator.TRAP_TYPE_NORMALIZE.get(warning['trap_type'])
                if normalized:
                    cursor.execute("UPDATE trap_library SET trap_type = ? WHERE id = ?", 
                                  (normalized, warning['id']))
                    result['fixed_count'] += 1
            
            # 4. 修复缺失
            cursor.execute("UPDATE word_frequency SET frequency_level = 'medium' WHERE frequency_level IS NULL OR frequency_level = ''")
            result['fixed_count'] += cursor.rowcount
            
            cursor.execute("UPDATE trap_library SET explanation = '请补充解释说明' WHERE explanation IS NULL OR explanation = ''")
            result['fixed_count'] += cursor.rowcount
            
            validator.conn.commit()
        
        else:
            result = {'success': False, 'error': f'未知的修复类型: {fix_type}'}
        
        validator.close()
        
        return jsonify(result)
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/kb/export/options')
def api_kb_export_options():
    """
    获取导出选项信息
    
    输出：可用的导出选项
    """
    try:
        exporter = KnowledgeBaseExporter(DB_PATH)
        options = exporter.get_export_options()
        exporter.close()
        
        return jsonify({
            'success': True,
            'options': options
        })
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/kb/export')
def api_kb_export():
    """
    导出知识库接口
    
    参数：
    - format: 导出格式 (xlsx, csv, json, md)
    - type: 导出类型 (vocab, trap, all)
    - pos_tag: 词性筛选
    - frequency_level: 频率等级筛选
    - trap_type: 陷阱类型筛选
    
    输出：导出的文件
    """
    try:
        # 获取参数
        export_format = request.args.get('format', 'xlsx')
        export_type = request.args.get('type', 'all')
        pos_tag = request.args.get('pos_tag', '')
        frequency_level = request.args.get('frequency_level', '')
        trap_type = request.args.get('trap_type', '')
        
        # 构建筛选条件
        filters = {}
        if pos_tag:
            filters['pos_tag'] = pos_tag
        if frequency_level:
            filters['frequency_level'] = frequency_level
        if trap_type:
            filters['trap_type'] = trap_type
        
        # 生成文件名
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        
        # 创建导出器
        exporter = KnowledgeBaseExporter(DB_PATH)
        
        if export_format == 'xlsx':
            output_filename = f'知识库_{timestamp}.xlsx'
            output_path = os.path.join(app.config['UPLOAD_FOLDER'], output_filename)
            exporter.export_to_excel(output_path, filters if filters else None, export_type)
            mimetype = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            
        elif export_format == 'csv':
            if export_type == 'all':
                # CSV导出多个文件时返回JSON
                vocab_path = os.path.join(app.config['UPLOAD_FOLDER'], f'词汇库_{timestamp}.csv')
                trap_path = os.path.join(app.config['UPLOAD_FOLDER'], f'陷阱库_{timestamp}.csv')
                
                exporter.export_to_csv(vocab_path, filters if filters else None, 'vocab')
                exporter.export_to_csv(trap_path, filters if filters else None, 'trap')
                
                return jsonify({
                    'success': True,
                    'files': {
                        'vocab': f'/api/kb/export/file?path={vocab_path}',
                        'trap': f'/api/kb/export/file?path={trap_path}'
                    }
                })
            else:
                output_filename = f'知识库_{timestamp}.csv'
                output_path = os.path.join(app.config['UPLOAD_FOLDER'], output_filename)
                exporter.export_to_csv(output_path, filters if filters else None, export_type)
            mimetype = 'text/csv'
            
        elif export_format == 'json':
            output_filename = f'知识库_{timestamp}.json'
            output_path = os.path.join(app.config['UPLOAD_FOLDER'], output_filename)
            exporter.export_to_json(output_path, filters if filters else None, export_type)
            mimetype = 'application/json'
            
        elif export_format == 'md':
            output_filename = f'知识库_{timestamp}.md'
            output_path = os.path.join(app.config['UPLOAD_FOLDER'], output_filename)
            exporter.export_to_markdown(output_path, filters if filters else None, export_type)
            mimetype = 'text/markdown'
            
        else:
            return jsonify({'success': False, 'error': '不支持的导出格式'}), 400
        
        exporter.close()
        
        return send_file(
            output_path,
            as_attachment=True,
            download_name=output_filename,
            mimetype=mimetype
        )
        
    except ImportError as e:
        if 'openpyxl' in str(e):
            return jsonify({'success': False, 'error': 'Excel导出需要安装openpyxl库，请运行: pip install openpyxl'}), 500
        raise
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/kb/export/file')
def api_kb_export_file():
    """
    获取导出的文件
    
    参数：
    - path: 文件路径
    """
    try:
        file_path = request.args.get('path', '')
        
        if not file_path or not os.path.exists(file_path):
            return jsonify({'success': False, 'error': '文件不存在'}), 404
        
        # 获取文件扩展名
        if file_path.endswith('.xlsx'):
            mimetype = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        elif file_path.endswith('.csv'):
            mimetype = 'text/csv'
        elif file_path.endswith('.json'):
            mimetype = 'application/json'
        elif file_path.endswith('.md'):
            mimetype = 'text/markdown'
        else:
            mimetype = 'application/octet-stream'
        
        filename = os.path.basename(file_path)
        
        return send_file(
            file_path,
            as_attachment=True,
            download_name=filename,
            mimetype=mimetype
        )
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/kb/export/all')
def api_kb_export_all():
    """
    打包导出所有知识库格式
    
    参数：
    - pos_tag: 词性筛选
    - frequency_level: 频率等级筛选
    - trap_type: 陷阱类型筛选
    
    输出：ZIP打包文件
    """
    try:
        # 获取筛选参数
        pos_tag = request.args.get('pos_tag', '')
        frequency_level = request.args.get('frequency_level', '')
        trap_type = request.args.get('trap_type', '')
        
        # 构建筛选条件
        filters = {}
        if pos_tag:
            filters['pos_tag'] = pos_tag
        if frequency_level:
            filters['frequency_level'] = frequency_level
        if trap_type:
            filters['trap_type'] = trap_type
        
        # 创建导出器
        exporter = KnowledgeBaseExporter(DB_PATH)
        
        # 导出所有格式
        exported_files = exporter.export_all(app.config['UPLOAD_FOLDER'], filters if filters else None)
        exporter.close()
        
        # 如果有ZIP文件，返回ZIP
        if 'zip' in exported_files and os.path.exists(exported_files['zip']):
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            zip_filename = f'知识库完整包_{timestamp}.zip'
            
            return send_file(
                exported_files['zip'],
                as_attachment=True,
                download_name=zip_filename,
                mimetype='application/zip'
            )
        
        # 否则返回文件列表
        return jsonify({
            'success': True,
            'files': exported_files
        })
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/kb/sync', methods=['POST'])
def api_kb_sync():
    """
    从Markdown文件同步到SQLite
    
    功能：从Markdown文件同步到SQLite
    """
    try:
        kb = get_kb_manager()
        kb.connect()
        
        synced_count = {'vocab': 0, 'trap': 0}
        
        # 遍历知识库数据目录
        for filename in os.listdir(KB_DATA_PATH):
            if not filename.endswith('.md'):
                continue
            
            filepath = os.path.join(KB_DATA_PATH, filename)
            
            with open(filepath, 'r', encoding='utf-8') as f:
                content = f.read()
            
            # 解析词汇库文件
            if '词汇库' in filename:
                synced_count['vocab'] += kb.sync_vocab_from_markdown(content)
            
            # 解析陷阱库文件
            elif '陷阱库' in filename:
                synced_count['trap'] += kb.sync_traps_from_markdown(content)
        
        kb.close()
        
        return jsonify({
            'success': True,
            'message': f'同步完成！词汇库更新 {synced_count["vocab"]} 条，陷阱库更新 {synced_count["trap"]} 条'
        })
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/kb/vocab')
def api_kb_vocab():
    """
    获取词汇列表（支持分页和筛选）
    
    参数：page, page_size, search, pos_filter, freq_filter
    输出：词汇列表及分页信息
    """
    try:
        page = request.args.get('page', 1, type=int)
        page_size = request.args.get('page_size', 100, type=int)
        search = request.args.get('search', '').strip()
        pos_filter = request.args.get('pos_filter', '').strip()
        freq_filter = request.args.get('freq_filter', '').strip()
        
        kb = get_kb_manager()
        kb.connect()
        
        all_vocab = kb.get_all_vocab()
        
        # 筛选
        filtered_vocab = all_vocab
        if search:
            filtered_vocab = [v for v in filtered_vocab if search.lower() in v.get('word', '').lower()]
        if pos_filter:
            filtered_vocab = [v for v in filtered_vocab if v.get('pos_tag') == pos_filter]
        if freq_filter:
            filtered_vocab = [v for v in filtered_vocab if v.get('frequency_level') == freq_filter]
        
        total = len(filtered_vocab)
        
        # 分页
        start = (page - 1) * page_size
        end = start + page_size
        paginated_vocab = filtered_vocab[start:end]
        
        kb.close()
        
        return jsonify({
            'success': True,
            'words': paginated_vocab,
            'total': total,
            'page': page,
            'page_size': page_size,
            'total_pages': (total + page_size - 1) // page_size if page_size > 0 else 0
        })
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/kb/vocab', methods=['POST'])
def api_kb_vocab_add():
    """
    添加词汇（支持例句字段）
    
    输入：word, frequency_level, pos_tag, example
    """
    try:
        data = request.get_json()
        
        word = data.get('word', '').strip()
        frequency_level = data.get('frequency_level', 'medium')
        pos_tag = data.get('pos_tag', '')
        example = data.get('example', '')
        
        if not word:
            return jsonify({'success': False, 'error': '词汇不能为空'}), 400
        
        kb = get_kb_manager()
        kb.connect()
        
        # 添加词汇
        kb.add_word_frequency(word, frequency_level, pos_tag)
        
        # 如果有例句，更新数据库
        if example:
            cursor = kb.conn.cursor()
            cursor.execute("""
                UPDATE word_frequency SET example = ? WHERE word = ?
            """, (example, word.lower()))
            kb.conn.commit()
        
        kb.close()
        
        return jsonify({
            'success': True,
            'message': f'词汇 {word} 添加成功'
        })
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/kb/vocab/<word>', methods=['PUT'])
def api_kb_vocab_update(word):
    """更新词汇字段"""
    try:
        data = request.get_json()
        field = data.get('field')
        value = data.get('value', '')
        
        kb = get_kb_manager()
        kb.connect()
        
        cursor = kb.conn.cursor()
        
        # 根据字段更新
        if field == 'word':
            cursor.execute("UPDATE word_frequency SET word = ? WHERE word = ?", 
                         (value.lower(), word.lower()))
        elif field == 'pos_tag':
            cursor.execute("UPDATE word_frequency SET pos_tag = ? WHERE word = ?", 
                         (value, word.lower()))
        elif field == 'frequency_level':
            cursor.execute("UPDATE word_frequency SET frequency_level = ? WHERE word = ?", 
                         (value, word.lower()))
        elif field == 'example':
            cursor.execute("UPDATE word_frequency SET example = ? WHERE word = ?", 
                         (value, word.lower()))
        
        kb.conn.commit()
        kb.close()
        
        return jsonify({
            'success': True,
            'message': '更新成功'
        })
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/kb/vocab/<word>', methods=['DELETE'])
def api_kb_vocab_delete(word):
    """删除词汇"""
    try:
        kb = get_kb_manager()
        kb.connect()
        kb.delete_vocab(word)
        kb.close()
        
        return jsonify({
            'success': True,
            'message': f'词汇 {word} 删除成功'
        })
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/kb/traps')
def api_kb_traps():
    """
    获取陷阱列表（支持分页和筛选）
    
    参数：page, page_size, search, type_filter
    输出：陷阱列表及分页信息
    """
    try:
        page = request.args.get('page', 1, type=int)
        page_size = request.args.get('page_size', 100, type=int)
        search = request.args.get('search', '').strip()
        type_filter = request.args.get('type_filter', '').strip()
        
        kb = get_kb_manager()
        kb.connect()
        
        all_traps = kb.get_all_traps()
        
        # 筛选
        filtered_traps = all_traps
        if search:
            filtered_traps = [t for t in filtered_traps if search.lower() in t.get('correct_word', '').lower()]
        if type_filter:
            filtered_traps = [t for t in filtered_traps if t.get('trap_type') == type_filter]
        
        total = len(filtered_traps)
        
        # 分页
        start = (page - 1) * page_size
        end = start + page_size
        paginated_traps = filtered_traps[start:end]
        
        kb.close()
        
        return jsonify({
            'success': True,
            'traps': paginated_traps,
            'total': total,
            'page': page,
            'page_size': page_size,
            'total_pages': (total + page_size - 1) // page_size if page_size > 0 else 0
        })
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/kb/traps', methods=['POST'])
def api_kb_traps_add():
    """
    添加陷阱
    
    功能：添加陷阱
    """
    try:
        data = request.get_json()
        
        correct_word = data.get('correct_word', '').strip()
        trap_type = data.get('trap_type', '词汇辨析')
        trap_words = data.get('trap_words', '')
        explanation = data.get('explanation', '')
        pos_tag = data.get('pos_tag', 'unknown')
        
        if not correct_word:
            return jsonify({'success': False, 'error': '正确词汇不能为空'}), 400
        
        if isinstance(trap_words, str):
            trap_words = [t.strip() for t in trap_words.split(',') if t.strip()]
        
        kb = get_kb_manager()
        kb.connect()
        kb.add_trap(correct_word, trap_type, trap_words, explanation, pos_tag)
        kb.close()
        
        return jsonify({
            'success': True,
            'message': f'陷阱词组 {correct_word} 添加成功'
        })
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/kb/traps/<word>', methods=['DELETE'])
def api_kb_traps_delete(word):
    """删除陷阱"""
    try:
        kb = get_kb_manager()
        kb.connect()
        kb.delete_trap(word)
        kb.close()
        
        return jsonify({
            'success': True,
            'message': f'陷阱词组 {word} 删除成功'
        })
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


# ==================== 批量导入API接口 ====================

@app.route('/api/kb/import', methods=['POST'])
def api_kb_import():
    """
    批量导入知识库
    
    功能：支持 Excel、CSV、JSON 格式文件导入
    输入：file (multipart/form-data)
    输出：导入结果
    """
    try:
        # 检查是否有文件
        if 'file' not in request.files:
            return jsonify({'success': False, 'error': '请选择要导入的文件'}), 400
        
        file = request.files['file']
        
        if file.filename == '':
            return jsonify({'success': False, 'error': '文件名为空'}), 400
        
        # 检查文件格式
        filename = file.filename.lower()
        if not (filename.endswith('.xlsx') or filename.endswith('.xls') or 
                filename.endswith('.csv') or filename.endswith('.json')):
            return jsonify({'success': False, 'error': '不支持的文件格式，请使用 xlsx/xls/csv/json 格式'}), 400
        
        # 创建上传目录
        import_dir = os.path.join('data', 'imports')
        os.makedirs(import_dir, exist_ok=True)
        
        # 保存上传的文件
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        safe_filename = f'{timestamp}_{file.filename}'
        file_path = os.path.join(import_dir, safe_filename)
        file.save(file_path)
        
        # 根据文件格式导入
        importer = KnowledgeBaseImporter(DB_PATH)
        
        try:
            if filename.endswith('.xlsx') or filename.endswith('.xls'):
                result = importer.import_from_excel(file_path)
            elif filename.endswith('.csv'):
                result = importer.import_from_csv(file_path)
            elif filename.endswith('.json'):
                result = importer.import_from_json(file_path)
        finally:
            importer.close()
        
        # 清理上传的文件
        try:
            os.remove(file_path)
        except:
            pass
        
        if result.success:
            return jsonify({
                'success': True,
                'message': result.message,
                'data': {
                    'total_rows': result.total_rows,
                    'imported': result.imported,
                    'duplicates': result.duplicates,
                    'errors': result.errors[:10] if result.errors else []  # 最多返回10条错误
                }
            })
        else:
            return jsonify({
                'success': False,
                'error': result.message
            }), 400
            
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/kb/import/template', methods=['GET'])
def api_kb_import_template():
    """
    下载导入模板
    
    功能：生成并下载 Excel 模板文件
    输入：type (vocab/trap)
    输出：Excel 文件
    """
    try:
        template_type = request.args.get('type', 'vocab')
        
        if template_type not in ['vocab', 'trap']:
            return jsonify({'success': False, 'error': '无效的模板类型'}), 400
        
        # 创建模板文件
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'导入模板_{template_type}_{timestamp}.xlsx'
        template_dir = os.path.join('data', 'templates')
        os.makedirs(template_dir, exist_ok=True)
        template_path = os.path.join(template_dir, filename)
        
        if create_template_excel(template_path, template_type):
            return send_file(
                template_path,
                as_attachment=True,
                download_name=filename,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
        else:
            return jsonify({'success': False, 'error': '创建模板失败'}), 500
            
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/kb/import/preview', methods=['POST'])
def api_kb_import_preview():
    """
    预览导入文件内容
    
    功能：在导入前预览文件数据
    输入：file (multipart/form-data)
    输出：预览数据（前10行）
    """
    try:
        if 'file' not in request.files:
            return jsonify({'success': False, 'error': '请选择要预览的文件'}), 400
        
        file = request.files['file']
        
        if file.filename == '':
            return jsonify({'success': False, 'error': '文件名为空'}), 400
        
        filename = file.filename.lower()
        if not (filename.endswith('.xlsx') or filename.endswith('.xls') or 
                filename.endswith('.csv') or filename.endswith('.json')):
            return jsonify({'success': False, 'error': '不支持的文件格式'}), 400
        
        # 保存临时文件
        import_dir = os.path.join('data', 'imports')
        os.makedirs(import_dir, exist_ok=True)
        temp_path = os.path.join(import_dir, f'temp_preview_{file.filename}')
        file.save(temp_path)
        
        try:
            # 根据文件格式读取预览数据
            if filename.endswith('.xlsx') or filename.endswith('.xls'):
                import openpyxl
                wb = openpyxl.load_workbook(temp_path, data_only=True)
                ws = wb.active
                
                # 读取表头和前10行数据
                headers = [str(cell.value) if cell.value else "" for cell in ws[1]]
                rows = []
                for row in ws.iter_rows(min_row=2, max_row=11, values_only=True):
                    if any(cell for cell in row):
                        rows.append([str(cell) if cell else "" for cell in row])
                
                wb.close()
                
            elif filename.endswith('.csv'):
                encodings = ['utf-8', 'gbk', 'gb2312', 'utf-8-sig']
                headers = []
                rows = []
                
                for enc in encodings:
                    try:
                        with open(temp_path, 'r', encoding=enc) as f:
                            reader = csv.reader(f)
                            headers = [cell.strip() for cell in next(reader)]
                            for i, row in enumerate(reader):
                                if i >= 10:
                                    break
                                if any(cell.strip() for cell in row):
                                    rows.append([cell.strip() for cell in row])
                        break
                    except UnicodeDecodeError:
                        continue
                        
            elif filename.endswith('.json'):
                with open(temp_path, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                
                if isinstance(data, list):
                    headers = list(data[0].keys()) if data and isinstance(data[0], dict) else []
                    rows = [[str(item.get(h, '')) for h in headers] for item in data[:10]]
                elif isinstance(data, dict):
                    key = next((k for k in ['vocab', 'words', 'traps', 'data'] if k in data), None)
                    if key:
                        items = data[key] if isinstance(data[key], list) else [data[key]]
                        headers = list(items[0].keys()) if items and isinstance(items[0], dict) else []
                        rows = [[str(item.get(h, '')) for h in headers] for item in items[:10]]
                    else:
                        headers = list(data.keys())[:10]
                        rows = [[str(data.get(h, '')) for h in headers]]
                else:
                    return jsonify({'success': False, 'error': 'JSON格式不正确'}), 400
            
            # 检测导入类型
            importer = KnowledgeBaseImporter(DB_PATH)
            import_type = importer._detect_import_type(headers)
            
            return jsonify({
                'success': True,
                'data': {
                    'headers': headers,
                    'rows': rows,
                    'import_type': import_type,
                    'total_hint': f'可导入约 {len(rows)} 条数据' if rows else '文件可能为空'
                }
            })
            
        finally:
            # 清理临时文件
            try:
                os.remove(temp_path)
            except:
                pass
            
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


# ==================== 备份恢复API接口 ====================

@app.route('/api/kb/backup', methods=['POST'])
def api_kb_backup():
    """
    创建知识库备份
    
    功能：手动创建备份，支持添加备注
    输入：note (可选)
    输出：备份结果
    """
    try:
        data = request.get_json() or {}
        note = data.get('note', '')
        
        backup_manager = get_backup_manager()
        result = backup_manager.manual_backup(note)
        
        if result['success']:
            return jsonify({
                'success': True,
                'message': result['message'],
                'data': {
                    'filename': result['filename'],
                    'size': result['formatted_size'],
                    'note': result.get('note', '')
                }
            })
        else:
            return jsonify({
                'success': False,
                'error': result['error']
            }), 500
            
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/kb/backups', methods=['GET'])
def api_kb_backups():
    """
    获取备份列表
    
    输出：备份历史列表
    """
    try:
        backup_manager = get_backup_manager()
        backups = backup_manager.list_backups()
        
        return jsonify({
            'success': True,
            'backups': backups,
            'count': len(backups)
        })
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/kb/restore/<backup_id>', methods=['POST'])
def api_kb_restore(backup_id):
    """
    恢复知识库备份
    
    功能：从指定备份恢复，恢复前自动备份当前版本
    输入：backup_id (备份文件名)
    输出：恢复结果
    """
    try:
        backup_manager = get_backup_manager()
        result = backup_manager.restore(backup_id)
        
        if result['success']:
            return jsonify({
                'success': True,
                'message': result['message'],
                'data': {
                    'pre_backup': result.get('pre_backup', ''),
                    'before_stats': result.get('before_stats', {}),
                    'after_stats': result.get('after_stats', {})
                }
            })
        else:
            return jsonify({
                'success': False,
                'error': result['error']
            }), 500
            
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/kb/compare', methods=['POST'])
def api_kb_compare():
    """
    对比两个备份版本的差异
    
    输入：backup1, backup2 (备份文件名)
    输出：差异对比结果
    """
    try:
        data = request.get_json()
        backup1 = data.get('backup1', '')
        backup2 = data.get('backup2', '')
        
        if not backup1 or not backup2:
            return jsonify({
                'success': False,
                'error': '请提供两个备份文件名'
            }), 400
        
        backup_manager = get_backup_manager()
        result = backup_manager.compare_versions(backup1, backup2)
        
        return jsonify(result)
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/kb/backup/<backup_id>', methods=['DELETE'])
def api_kb_backup_delete(backup_id):
    """
    删除指定备份
    
    输入：backup_id (备份文件名)
    输出：删除结果
    """
    try:
        backup_manager = get_backup_manager()
        result = backup_manager.delete_backup(backup_id)
        
        return jsonify(result)
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/history/<task_id>')
def api_history_detail(task_id):
    """获取历史记录详情"""
    try:
        history_path = os.path.join(app.config['UPLOAD_FOLDER'], f'{task_id}.json')
        
        if not os.path.exists(history_path):
            return jsonify({'success': False, 'error': '记录不存在'}), 404
        
        with open(history_path, 'r', encoding='utf-8') as f:
            data = json.load(f)
        
        return jsonify({
            'success': True,
            'data': data
        })
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/history/<task_id>', methods=['DELETE'])
def api_history_delete(task_id):
    """删除历史记录"""
    try:
        history_path = os.path.join(app.config['UPLOAD_FOLDER'], f'{task_id}.json')
        
        if os.path.exists(history_path):
            os.remove(history_path)
        
        return jsonify({
            'success': True,
            'message': '删除成功'
        })
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


# ==================== 知识库统计API ====================

@app.route('/api/kb/statistics')
def api_kb_statistics():
    """
    获取知识库完整统计数据
    
    输出：包含覆盖率、分布、质量指标等完整统计
    """
    try:
        stats = KnowledgeBaseStatistics(DB_PATH)
        report = stats.generate_report()
        stats.close()
        
        return jsonify({
            'success': True,
            'data': report
        })
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/kb/statistics/overall')
def api_kb_statistics_overall():
    """
    获取整体统计数据
    
    输出：总条目数、数据库大小等
    """
    try:
        stats = KnowledgeBaseStatistics(DB_PATH)
        data = stats.get_overall_stats()
        stats.close()
        
        return jsonify({
            'success': True,
            'data': data
        })
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/kb/statistics/coverage')
def api_kb_statistics_coverage():
    """
    获取覆盖率分析
    
    输出：中考词汇覆盖率、词性分布、陷阱类型分布等
    """
    try:
        stats = KnowledgeBaseStatistics(DB_PATH)
        data = stats.get_coverage_analysis()
        stats.close()
        
        return jsonify({
            'success': True,
            'data': data
        })
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/kb/statistics/distribution')
def api_kb_statistics_distribution():
    """
    获取考点分布统计
    
    输出：词性分布饼图、频率等级分布、难度分布等
    """
    try:
        stats = KnowledgeBaseStatistics(DB_PATH)
        data = stats.get_distribution_stats()
        stats.close()
        
        return jsonify({
            'success': True,
            'data': data
        })
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/kb/statistics/quality')
def api_kb_statistics_quality():
    """
    获取数据质量指标
    
    输出：完整率、有效率、更新频率等
    """
    try:
        stats = KnowledgeBaseStatistics(DB_PATH)
        data = stats.get_quality_metrics()
        stats.close()
        
        return jsonify({
            'success': True,
            'data': data
        })
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


# ==================== 模板下载 ====================

@app.route('/templates/<filename>')
def download_template(filename):
    """
    下载导入模板文件
    
    参数：filename - 模板文件名
    """
    try:
        template_path = os.path.join('data/templates', filename)
        if not os.path.exists(template_path):
            return jsonify({'success': False, 'error': '模板文件不存在'}), 404
        
        return send_file(
            template_path,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/kb/import/vocab', methods=['POST'])
def api_kb_import_vocab():
    """
    批量导入词汇（从Excel/CSV）
    
    输入：file (multipart/form-data)
    """
    try:
        if 'file' not in request.files:
            return jsonify({'success': False, 'error': '请上传文件'}), 400
        
        file = request.files['file']
        if not file.filename.endswith(('.xlsx', '.csv')):
            return jsonify({'success': False, 'error': '仅支持 .xlsx 或 .csv 格式'}), 400
        
        import pandas as pd
        from io import BytesIO
        
        # 读取文件
        if file.filename.endswith('.xlsx'):
            df = pd.read_excel(file)
        else:
            df = pd.read_csv(file)
        
        # 验证列
        required_cols = ['单词']
        missing_cols = [col for col in required_cols if col not in df.columns]
        if missing_cols:
            return jsonify({'success': False, 'error': f'缺少必需列: {", ".join(missing_cols)}'}), 400
        
        kb = get_kb_manager()
        kb.connect()
        
        imported_count = 0
        for _, row in df.iterrows():
            word = str(row.get('单词', '')).strip()
            if not word or word == 'nan':
                continue
            
            frequency_level = str(row.get('频率等级', 'medium')).strip().lower()
            if frequency_level not in ['high', 'medium', 'low']:
                frequency_level = 'medium'
            
            pos_tag = str(row.get('词性', '')).strip()
            if pos_tag == 'nan':
                pos_tag = ''
            
            example = str(row.get('例句', '')).strip()
            if example == 'nan':
                example = ''
            
            kb.add_word_frequency(word, frequency_level, pos_tag)
            
            if example:
                cursor = kb.conn.cursor()
                cursor.execute("UPDATE word_frequency SET example = ? WHERE word = ?",
                             (example, word.lower()))
                kb.conn.commit()
            
            imported_count += 1
        
        kb.close()
        
        return jsonify({
            'success': True,
            'message': f'成功导入 {imported_count} 条词汇'
        })
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/kb/import/traps', methods=['POST'])
def api_kb_import_traps():
    """
    批量导入陷阱（从Excel/CSV）
    
    输入：file (multipart/form-data)
    """
    try:
        if 'file' not in request.files:
            return jsonify({'success': False, 'error': '请上传文件'}), 400
        
        file = request.files['file']
        if not file.filename.endswith(('.xlsx', '.csv')):
            return jsonify({'success': False, 'error': '仅支持 .xlsx 或 .csv 格式'}), 400
        
        import pandas as pd
        
        # 读取文件
        if file.filename.endswith('.xlsx'):
            df = pd.read_excel(file)
        else:
            df = pd.read_csv(file)
        
        # 验证列
        required_cols = ['正确答案', '陷阱类型']
        missing_cols = [col for col in required_cols if col not in df.columns]
        if missing_cols:
            return jsonify({'success': False, 'error': f'缺少必需列: {", ".join(missing_cols)}'}), 400
        
        kb = get_kb_manager()
        kb.connect()
        
        imported_count = 0
        for _, row in df.iterrows():
            correct_word = str(row.get('正确答案', '')).strip()
            if not correct_word or correct_word == 'nan':
                continue
            
            trap_type = str(row.get('陷阱类型', '词汇辨析')).strip()
            
            # 收集干扰项
            trap_words = []
            for col in ['干扰项1', '干扰项2', '干扰项3']:
                if col in df.columns:
                    trap_word = str(row.get(col, '')).strip()
                    if trap_word and trap_word != 'nan':
                        trap_words.append(trap_word)
            
            explanation = str(row.get('解析', '')).strip()
            if explanation == 'nan':
                explanation = ''
            
            pos_tag = str(row.get('词性', '')).strip()
            if pos_tag == 'nan':
                pos_tag = ''
            
            if trap_words:
                kb.add_trap(correct_word, trap_type, trap_words, explanation, pos_tag)
                imported_count += 1
        
        kb.close()
        
        return jsonify({
            'success': True,
            'message': f'成功导入 {imported_count} 条陷阱'
        })
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


# ==================== 文件上传API ====================

@app.route('/api/upload/image', methods=['POST'])
def api_upload_image():
    """
    上传图片并OCR识别
    
    输入：image文件 (jpg/png等)
    输出：识别后的文本内容
    """
    try:
        if 'file' not in request.files:
            return jsonify({'success': False, 'error': '没有上传文件'}), 400
        
        file = request.files['file']
        
        if file.filename == '':
            return jsonify({'success': False, 'error': '未选择文件'}), 400
        
        # 验证文件
        validate_file(file, max_size_mb=10)
        
        # 处理图片OCR
        text = process_uploaded_file(file, file_type='image')
        
        # 检查识别结果
        if not text or len(text.strip()) < 10:
            return jsonify({
                'success': False, 
                'error': '图片中未识别到足够的英文文字，请确保图片清晰且包含英文内容'
            }), 400
        
        # 统计信息
        word_count = len(text.split())
        
        return jsonify({
            'success': True,
            'text': text,
            'word_count': word_count,
            'message': f'成功识别 {word_count} 个英文单词'
        })
        
    except ImportError as e:
        return jsonify({
            'success': False, 
            'error': f'缺少依赖库: {str(e)}',
            'install_hint': '请运行: pip install pytesseract Pillow && 在系统中安装Tesseract OCR'
        }), 500
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/upload/file', methods=['POST'])
def api_upload_file():
    """
    上传文件并解析文本
    
    支持格式：PDF, DOCX, DOC
    输入：file文件
    输出：解析后的文本内容
    """
    try:
        if 'file' not in request.files:
            return jsonify({'success': False, 'error': '没有上传文件'}), 400
        
        file = request.files['file']
        
        if file.filename == '':
            return jsonify({'success': False, 'error': '未选择文件'}), 400
        
        # 验证文件
        validate_file(file, max_size_mb=10)
        
        # 获取文件信息
        file_info = get_file_info(file)
        
        # 根据扩展名确定文件类型
        ext = file_info['extension']
        if ext == '.pdf':
            file_type = 'pdf'
        elif ext == '.docx':
            file_type = 'docx'
        elif ext == '.doc':
            file_type = 'doc'
        else:
            return jsonify({'success': False, 'error': f'不支持的文件格式: {ext}'}), 400
        
        # 处理文件解析
        text = process_uploaded_file(file, file_type=file_type)
        
        # 检查解析结果
        if not text or len(text.strip()) < 50:
            return jsonify({
                'success': False, 
                'error': '文件中未识别到足够的文本内容'
            }), 400
        
        # 统计信息
        word_count = len(text.split())
        
        return jsonify({
            'success': True,
            'text': text,
            'word_count': word_count,
            'file_type': file_type,
            'filename': file_info['filename'],
            'message': f'成功解析 {word_count} 个英文单词'
        })
        
    except ImportError as e:
        return jsonify({
            'success': False, 
            'error': str(e)
        }), 500
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


# ==================== 素材工具箱API ====================

from core.article_fetcher import ArticleFetcher
from core.article_generator import ArticleGenerator, generate_article
from core.text_adjuster import TextAdjuster, adjust_text, analyze_text

# 初始化工具
article_fetcher = ArticleFetcher('data/article_sources.json')
article_generator = ArticleGenerator()
text_adjuster = TextAdjuster()


@app.route('/api/article/search', methods=['POST'])
def api_article_search():
    """搜索文章"""
    try:
        data = request.get_json()
        keyword = data.get('keyword', '')
        source_ids = data.get('source_ids')
        
        if not keyword:
            return jsonify({'success': False, 'error': '请输入搜索关键词'}), 400
        
        articles = article_fetcher.search_articles(keyword, source_ids)
        
        return jsonify({
            'success': True,
            'articles': articles,
            'count': len(articles)
        })
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/article/fetch', methods=['POST'])
def api_article_fetch():
    """获取文章内容"""
    try:
        data = request.get_json()
        url = data.get('url', '')
        
        if not url:
            return jsonify({'success': False, 'error': '请提供文章URL'}), 400
        
        article = article_fetcher.fetch_article_content(url)
        
        if article:
            return jsonify({
                'success': True,
                'article': article
            })
        else:
            return jsonify({'success': False, 'error': '无法获取文章内容'}), 404
            
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/article/generate', methods=['POST'])
def api_article_generate():
    """AI生成文章"""
    try:
        data = request.get_json()
        topic = data.get('topic', 'school')
        genre = data.get('genre', '记叙文')
        word_count = data.get('word_count', 150)
        grade_level = data.get('grade_level', '初二')
        
        article = article_generator.generate_article(
            topic=topic,
            genre=genre,
            word_count=word_count,
            grade_level=grade_level
        )
        
        return jsonify({
            'success': True,
            'article': article
        })
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/article/analyze', methods=['POST'])
def api_article_analyze():
    """分析文本难度"""
    try:
        data = request.get_json()
        text = data.get('text', '')
        
        if not text:
            return jsonify({'success': False, 'error': '请提供文本内容'}), 400
        
        analysis = text_adjuster.analyze_difficulty(text)
        
        return jsonify({
            'success': True,
            'analysis': analysis
        })
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/article/adjust', methods=['POST'])
def api_article_adjust():
    """调整文本难度"""
    try:
        data = request.get_json()
        text = data.get('text', '')
        target_level = data.get('target_level', '初二')
        
        if not text:
            return jsonify({'success': False, 'error': '请提供文本内容'}), 400
        
        result = text_adjuster.adjust_difficulty(text, target_level)
        
        return jsonify({
            'success': True,
            'result': result
        })
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/sources', methods=['GET'])
def api_get_sources():
    """获取文章来源列表"""
    try:
        sources = article_fetcher.get_sources()
        return jsonify({
            'success': True,
            'sources': sources
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/sources', methods=['POST'])
def api_add_source():
    """添加文章来源"""
    try:
        data = request.get_json()
        source = {
            'id': data.get('id', str(len(article_fetcher.sources) + 1)),
            'name': data.get('name', ''),
            'url': data.get('url', ''),
            'type': data.get('type', 'website'),
            'enabled': data.get('enabled', True),
            'description': data.get('description', '')
        }
        
        article_fetcher.sources.append(source)
        article_fetcher.save_sources(article_fetcher.sources)
        
        return jsonify({
            'success': True,
            'source': source
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


# ==================== 轻量化工具API ====================

@app.route('/api/tools/discussion', methods=['POST'])
def api_discussion_questions():
    """
    讨论问题生成接口
    
    输入：topic, num_questions, difficulty
    输出：讨论问题JSON
    """
    try:
        data = request.get_json()
        topic = data.get('topic', '').strip()
        num_questions = data.get('num_questions', 8)
        difficulty = data.get('difficulty', 'intermediate')
        
        if not topic:
            return jsonify({'success': False, 'error': '请输入主题关键词'}), 400
        
        result = generate_discussion_questions(topic, num_questions, difficulty)
        return jsonify(result)
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/tools/sentences', methods=['POST'])
def api_sentence_generation():
    """
    词汇造句接口
    
    输入：vocabulary, num_examples, difficulty
    输出：例句JSON
    """
    try:
        data = request.get_json()
        vocabulary = data.get('vocabulary', '').strip()
        num_examples = data.get('num_examples', 3)
        difficulty = data.get('difficulty', 'intermediate')
        
        if not vocabulary:
            return jsonify({'success': False, 'error': '请输入词汇列表'}), 400
        
        result = generate_sentences(vocabulary, num_examples, difficulty)
        return jsonify(result)
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/tools/dialogue', methods=['POST'])
def api_dialogue_generation():
    """
    对话生成接口
    
    输入：scenario, num_exchanges, difficulty
    输出：对话JSON
    """
    try:
        data = request.get_json()
        scenario = data.get('scenario', '').strip()
        num_exchanges = data.get('num_exchanges', 8)
        difficulty = data.get('difficulty', 'intermediate')
        
        if not scenario:
            return jsonify({'success': False, 'error': '请输入场景描述'}), 400
        
        result = generate_dialogue(scenario, num_exchanges, difficulty)
        return jsonify(result)
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/tools/error-correction', methods=['POST'])
def api_error_correction():
    """
    改错题生成接口
    
    输入：grammar_point, num_sentences, difficulty
    输出：改错题JSON
    """
    try:
        data = request.get_json()
        grammar_point = data.get('grammar_point', '').strip()
        num_sentences = data.get('num_sentences', 8)
        difficulty = data.get('difficulty', 'intermediate')
        
        if not grammar_point:
            return jsonify({'success': False, 'error': '请输入语法点'}), 400
        
        result = generate_error_correction(grammar_point, num_sentences, difficulty)
        return jsonify(result)
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/tools/essential-vocabulary', methods=['POST'])
def api_essential_vocabulary():
    """
    核心词汇提取接口
    
    输入：topic, num_vocab, difficulty
    输出：核心词汇JSON
    """
    try:
        data = request.get_json()
        topic = data.get('topic', '').strip()
        num_vocab = data.get('num_vocab', 15)
        difficulty = data.get('difficulty', 'intermediate')
        
        if not topic:
            return jsonify({'success': False, 'error': '请输入主题关键词'}), 400
        
        result = generate_essential_vocabulary(topic, num_vocab, difficulty)
        return jsonify(result)
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/tools/pros-cons', methods=['POST'])
def api_pros_cons():
    """
    利弊分析接口
    
    输入：topic, num_points
    输出：利弊分析JSON
    """
    try:
        data = request.get_json()
        topic = data.get('topic', '').strip()
        num_points = data.get('num_points', 5)
        
        if not topic:
            return jsonify({'success': False, 'error': '请输入主题'}), 400
        
        result = generate_pros_cons(topic, num_points)
        return jsonify(result)
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/tools/famous-quotes', methods=['POST'])
def api_famous_quotes():
    """
    名人名言接口
    
    输入：topic, person, num_quotes
    输出：名言JSON
    """
    try:
        data = request.get_json()
        topic = data.get('topic', '').strip() or None
        person = data.get('person', '').strip() or None
        num_quotes = data.get('num_quotes', 6)
        
        if not topic and not person:
            return jsonify({'success': False, 'error': '请输入主题或名人名字'}), 400
        
        result = generate_famous_quotes(topic, person, num_quotes)
        return jsonify(result)
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/tools/essay-topics', methods=['POST'])
def api_essay_topics():
    """
    作文题目接口
    
    输入：topic, num_topics, include_types
    输出：作文题目JSON
    """
    try:
        data = request.get_json()
        topic = data.get('topic', '').strip()
        num_topics = data.get('num_topics', 10)
        include_types = data.get('include_types', 'all')
        
        if not topic:
            return jsonify({'success': False, 'error': '请输入主题领域'}), 400
        
        result = generate_essay_topics(topic, num_topics, include_types)
        return jsonify(result)
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/tools/interesting-facts', methods=['POST'])
def api_interesting_facts():
    """
    趣味知识接口
    
    输入：topic, num_facts
    输出：趣味知识JSON
    """
    try:
        data = request.get_json()
        topic = data.get('topic', '').strip()
        num_facts = data.get('num_facts', 10)
        
        if not topic:
            return jsonify({'success': False, 'error': '请输入主题'}), 400
        
        result = generate_interesting_facts(topic, num_facts)
        return jsonify(result)
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/tools/three-titles', methods=['POST'])
def api_three_titles():
    """
    标题选择接口
    
    输入：content, num_options
    输出：标题选项JSON
    """
    try:
        data = request.get_json()
        content = data.get('content', '').strip()
        num_options = data.get('num_options', 3)
        
        if not content:
            return jsonify({'success': False, 'error': '请输入短文内容'}), 400
        
        result = generate_three_titles(content, num_options)
        return jsonify(result)
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/tools/creative-writing', methods=['POST'])
def api_creative_writing():
    """
    创意写作接口
    
    输入：topic, vocabulary, num_tasks
    输出：创意写作任务JSON
    """
    try:
        data = request.get_json()
        topic = data.get('topic', '').strip() or None
        vocabulary = data.get('vocabulary', '').strip() or None
        num_tasks = data.get('num_tasks', 4)
        
        if not topic and not vocabulary:
            return jsonify({'success': False, 'error': '请输入主题或词汇'}), 400
        
        result = generate_creative_writing(topic, vocabulary, num_tasks)
        return jsonify(result)
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/tools/odd-one-out', methods=['POST'])
def api_odd_one_out():
    """
    找不同接口
    
    输入：topic, num_groups, difficulty
    输出：找不同练习JSON
    """
    try:
        data = request.get_json()
        topic = data.get('topic', '').strip() or None
        num_groups = data.get('num_groups', 6)
        difficulty = data.get('difficulty', 'intermediate')
        
        result = generate_odd_one_out(topic, num_groups, difficulty)
        return jsonify(result)
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


# ==================== 错误处理 ====================

@app.errorhandler(404)
def not_found(e):
    return jsonify({'success': False, 'error': '页面不存在'}), 404


@app.errorhandler(500)
def server_error(e):
    return jsonify({'success': False, 'error': '服务器内部错误'}), 500


# ==================== CEFR文本分析API ====================

# 初始化CEFR分析器（已在文件开头导入）
cefr_analyzer = CEFRAnalyzer()


@app.route('/cefr-test')
def cefr_test():
    """
    CEFR文本难度分析测试页面
    独立模块，暂不集成到主页
    """
    return render_template('cefr_test.html')


@app.route('/api/cefr/analyze', methods=['POST'])
def api_cefr_analyze():
    """
    CEFR文本分析接口
    
    输入：text - 待分析的英文文本
    输出：分析结果JSON
    """
    try:
        data = request.get_json()
        text = data.get('text', '').strip()
        
        if not text:
            return jsonify({'success': False, 'error': '请输入文本内容'}), 400
        
        result = cefr_analyzer.analyze(text)
        
        return jsonify({
            'success': True,
            'result': result
        })
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


# ==================== 启动应用 ====================

if __name__ == '__main__':
    # 初始化数据库（如果不存在）
    if not os.path.exists(DB_PATH):
        from core.kb_manager import init_database
        init_database(DB_PATH)
        print(f"已初始化数据库: {DB_PATH}")
    
    # 获取端口（Railway会自动分配$PORT）
    port = int(os.environ.get('PORT', 5000))
    
    # 启动服务器
    print("=" * 50)
    print("智能出题引擎 Web 版")
    print("=" * 50)
    print(f"访问地址: http://localhost:{port}")
    print("按 Ctrl+C 停止服务")
    print("=" * 50)
    
    app.run(debug=False, host='0.0.0.0', port=port)
